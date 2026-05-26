"""
板块资金趋势追踪模块
读取 board_amount_ratio 表，计算5/10/20日资金趋势，输出报告/Excel/JSON
运行：python -m analysis.board_trend_tracker [--date YYYYMMDD]
"""
import argparse
import json
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import psycopg2
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from data.config import DATABASE_DSN
from analysis.board_alias import aggregate_by_display_name

logger = logging.getLogger(__name__)

REPORTS_DIR = Path(__file__).resolve().parents[1] / "reports" / "daily"


# ── 格式化辅助 ──

def _to_float(v):
    try:
        if v is None or pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


def _fmt_ratio_pct(v):
    x = _to_float(v)
    return "-" if x is None else f"{x * 100:.2f}%"


def _fmt_change_pct(v):
    x = _to_float(v)
    return "-" if x is None else f"{x * 100:+.2f}%"


def _fmt_num(v, digits=1):
    x = _to_float(v)
    return "-" if x is None else f"{x:.{digits}f}"


def _excel_num(v, mult=1):
    x = _to_float(v)
    return None if x is None else x * mult


def _get_db_conn():
    if not DATABASE_DSN:
        return None
    try:
        return psycopg2.connect(DATABASE_DSN)
    except Exception as e:
        print(f"[错误] 数据库连接失败：{e}")
        return None


def _read_board_data(trade_date, lookback=60):
    """读取 board_amount_ratio 最近 N 天数据"""
    conn = _get_db_conn()
    if conn is None:
        return None
    start = (datetime.strptime(trade_date, "%Y%m%d") - timedelta(days=lookback)).strftime("%Y-%m-%d")
    end_dt = datetime.strptime(trade_date, "%Y%m%d")
    df = pd.read_sql(
        "SELECT * FROM board_amount_ratio WHERE trade_date >= %s AND trade_date <= %s ORDER BY trade_date",
        conn, params=(start, end_dt)
    )
    conn.close()
    if df.empty:
        return None
    df["trade_date"] = pd.to_datetime(df["trade_date"])
    for c in ["pct_chg", "amount", "amount_ratio", "turnover", "up_count", "down_count"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def _calc_metrics(df, trade_date, windows):
    """计算每个板块的核心指标"""
    end_dt = datetime.strptime(trade_date, "%Y%m%d")
    results = []

    for board_type in ["行业", "概念"]:
        bdf = df[df["board_type"] == board_type].copy()
        if bdf.empty:
            continue
        for board_name, group in bdf.groupby("board_name"):
            group = group.sort_values("trade_date")
            latest = group.iloc[-1]
            dates = sorted(group["trade_date"].unique())

            metrics = {
                "board_name": board_name,
                "board_type": board_type,
                "board_code": str(latest.get("board_code", "")),
                "latest_pct_chg": latest.get("pct_chg"),
                "latest_amount": latest.get("amount"),
                "latest_amount_ratio": latest.get("amount_ratio"),
                "latest_turnover": latest.get("turnover"),
                "latest_up_count": latest.get("up_count"),
                "latest_down_count": latest.get("down_count"),
                "leader_name": str(latest.get("leader_name", "")),
                "leader_pct_chg": latest.get("leader_pct_chg"),
                "trade_days": len(dates),
            }

            # 上涨比例
            uc = metrics["latest_up_count"] or 0
            dc = metrics["latest_down_count"] or 0
            metrics["up_ratio"] = uc / (uc + dc) if (uc + dc) > 0 else None

            # 窗口指标
            for w in windows:
                window = group[group["trade_date"] >= end_dt - timedelta(days=w * 2)].tail(w)
                if len(window) < w:
                    metrics[f"amount_ratio_change_{w}d"] = None
                    metrics[f"amount_ratio_mean_{w}d"] = None
                    metrics[f"amount_vs_{w}d_mean"] = None
                    metrics[f"pct_positive_days_{w}d"] = None
                    continue

                ratios = window["amount_ratio"].dropna()
                amounts = window["amount"].dropna()
                pct_chgs = window["pct_chg"].dropna()

                if len(ratios) >= 1:
                    first_ratio = ratios.iloc[0]
                    last_ratio = ratios.iloc[-1]
                    metrics[f"amount_ratio_change_{w}d"] = last_ratio - first_ratio if pd.notna(first_ratio) else None
                    metrics[f"amount_ratio_mean_{w}d"] = ratios.mean()

                if len(amounts) >= 2 and metrics["latest_amount"]:
                    mean_amt = amounts.mean()
                    if mean_amt and mean_amt > 0:
                        metrics[f"amount_vs_{w}d_mean"] = metrics["latest_amount"] / mean_amt
                    else:
                        metrics[f"amount_vs_{w}d_mean"] = None

                metrics[f"pct_positive_days_{w}d"] = int((pct_chgs > 0).sum()) if len(pct_chgs) > 0 else 0

            # 连续上升天数
            streak = 0
            ratios_list = group["amount_ratio"].dropna().values
            if len(ratios_list) >= 2:
                i = len(ratios_list) - 1
                while i > 0 and streak < 5:
                    if ratios_list[i] > ratios_list[i - 1]:
                        streak += 1
                        i -= 1
                    else:
                        break
            metrics["amount_ratio_up_streak"] = streak

            results.append(metrics)

    return pd.DataFrame(results)


def _classify_flow(row):
    """根据指标判断资金状态"""
    c5 = row.get("amount_ratio_change_5d")
    c10 = row.get("amount_ratio_change_10d")
    c20 = row.get("amount_ratio_change_20d")
    vs5 = row.get("amount_vs_5d_mean")
    pct = row.get("latest_pct_chg") or 0
    up_ratio = row.get("up_ratio") or 0
    streak = row.get("amount_ratio_up_streak") or 0
    mean_20 = row.get("amount_ratio_mean_20d")

    # 持续流入
    if c5 is not None and c5 > 0 and streak >= 2 and pct > 0 and up_ratio >= 0.55:
        return "持续流入"

    # 加速流入
    if c5 is not None and c5 > 0.002 and vs5 is not None and vs5 >= 1.3 and pct > 1:
        return "加速流入"

    # 资金回流
    if c5 is not None and c5 > 0 and c10 is not None and c10 <= 0 and pct > 0:
        return "资金回流"

    # 高位分歧
    if mean_20 is not None and row.get("latest_amount_ratio") and row["latest_amount_ratio"] >= mean_20 \
            and vs5 is not None and vs5 >= 1.3 and (pct <= 1 or up_ratio < 0.5):
        return "高位分歧"

    # 资金退潮
    if c5 is not None and c5 < 0 and pct < 0:
        return "资金退潮"
    p5 = row.get("pct_positive_days_5d") or 0
    if c10 is not None and c10 < 0 and p5 <= 2:
        return "资金退潮"

    return "无明显方向"


def _classify_lifecycle(row):
    """按优先级判断板块生命周期：分歧→退潮→加速→主升→修复→启动→沉寂"""
    fs = row.get("flow_status", "")
    ts = row.get("trend_score", 0)
    c5 = row.get("amount_ratio_change_5d")
    c10 = row.get("amount_ratio_change_10d")
    vs5 = row.get("amount_vs_5d_mean")
    ar20 = row.get("amount_ratio_mean_20d")
    ar_latest = row.get("latest_amount_ratio")
    pct = row.get("latest_pct_chg") or 0
    ur = row.get("up_ratio") or 0

    # 分歧期
    if fs == "高位分歧" or (ar_latest is not None and ar20 is not None and ar_latest >= ar20
            and vs5 is not None and vs5 >= 1.2 and (pct <= 1 or ur < 0.5)):
        return "分歧期"
    # 退潮期
    if fs == "资金退潮" or (c5 is not None and c5 < 0 and pct < 0):
        return "退潮期"
    # 加速期
    if fs == "加速流入" or (c5 is not None and c5 > 0.002 and vs5 is not None and vs5 >= 1.3 and pct > 1):
        return "加速期"
    # 主升期
    if ts >= 70 and c10 is not None and c10 > 0 and c5 is not None and c5 > 0 and ur >= 0.55 and pct > 0:
        return "主升期"
    # 修复期
    if fs == "资金回流" or (c5 is not None and c5 > 0 and c10 is not None and c10 <= 0 and pct > 0):
        return "修复期"
    # 启动期
    if c5 is not None and c5 > 0 and ts >= 40 and ts < 70 and pct > 0:
        return "启动期"
    return "沉寂期"


def _lifecycle_desc(stage):
    return {
        "启动期": "资金开始关注，持续性仍需观察。",
        "加速期": "成交占比快速提升，短线资金明显加速。",
        "主升期": "中短期资金持续流入，主线确认度较高。",
        "分歧期": "成交活跃但上涨扩散不足，需警惕高位分歧。",
        "修复期": "前期退潮后资金回流，观察修复持续性。",
        "退潮期": "成交占比下降且表现转弱，资金热度下降。",
        "沉寂期": "资金方向不明显，暂时不是市场重点。",
    }.get(stage, "")


def _classify_lifecycle_signal(prev_stage, curr_stage):
    if not prev_stage:
        return "新增观察"
    positive = {
        ("沉寂期", "启动期"): "新启动",
        ("启动期", "加速期"): "主线加强",
        ("启动期", "主升期"): "主线确认",
        ("修复期", "启动期"): "修复转强",
        ("修复期", "加速期"): "资金回流加强",
        ("加速期", "主升期"): "主线延续",
        ("主升期", "主升期"): "主线延续",
    }
    warning = {
        ("主升期", "分歧期"): "主线分歧",
        ("加速期", "分歧期"): "加速后分歧",
        ("分歧期", "退潮期"): "分歧转退潮",
        ("主升期", "退潮期"): "主线退潮",
        ("启动期", "退潮期"): "启动失败",
        ("修复期", "退潮期"): "修复失败",
    }
    if (prev_stage, curr_stage) in positive:
        return positive[(prev_stage, curr_stage)]
    if (prev_stage, curr_stage) in warning:
        return warning[(prev_stage, curr_stage)]
    if curr_stage == prev_stage:
        return "状态延续"
    if curr_stage in ("启动期", "加速期", "主升期"):
        return "状态转强"
    if curr_stage in ("分歧期", "退潮期"):
        return "状态转弱"
    return "状态变化"


def _calc_mainline_streak(row):
    """主线连续天数近似：连续上升+上涨天数"""
    streak = row.get("amount_ratio_up_streak") or 0
    p5 = row.get("pct_positive_days_5d") or 0
    stage = row.get("life_cycle", "")
    if stage in ("启动期", "加速期", "主升期", "修复期"):
        return min(streak + p5, 5)
    return 0


def _calc_trend_score(row):
    """计算趋势评分 0-100"""
    score = 0
    c5 = row.get("amount_ratio_change_5d")
    c10 = row.get("amount_ratio_change_10d")
    c20 = row.get("amount_ratio_change_20d")
    streak = row.get("amount_ratio_up_streak") or 0
    pct = row.get("latest_pct_chg") or 0
    up_ratio = row.get("up_ratio") or 0
    vs5 = row.get("amount_vs_5d_mean")
    vs10 = row.get("amount_vs_10d_mean")

    # 资金维度 max 40
    if c5 is not None and c5 > 0:
        score += 10
    if c10 is not None and c10 > 0:
        score += 10
    if c20 is not None and c20 > 0:
        score += 10
    if streak >= 2:
        score += 10

    # 价格维度 max 20
    if pct > 0:
        score += 6
    if pct > 1:
        score += 8
    p5 = row.get("pct_positive_days_5d") or 0
    if p5 >= 3:
        score += 6

    # 扩散维度 max 20
    if up_ratio >= 0.55:
        score += 8
    if up_ratio >= 0.65:
        score += 6
    if up_ratio >= 0.55:
        score += 6

    # 活跃维度 max 20
    if vs5 is not None and vs5 > 1:
        score += 8
    if vs10 is not None and vs10 > 1:
        score += 6
    if row.get("latest_turnover") and row["latest_turnover"]:
        score += 6

    return min(score, 100)


def _score_to_label(score):
    if score >= 80:
        return "强主线"
    elif score >= 60:
        return "潜在主线"
    elif score >= 40:
        return "短线热点"
    elif score >= 20:
        return "弱跟踪"
    return "非主线"


def _generate_markdown(df, trade_date, windows_available):
    """生成 Markdown 报告"""
    date_display = f"{trade_date[:4]}-{trade_date[4:6]}-{trade_date[6:]}"
    lines = [f"# 板块资金趋势追踪报告 | {date_display}", ""]

    # 一、总体结论
    lines.append("## 一、总体结论")
    lines.append("")

    inflow = df[df["flow_status"].isin(["持续流入", "加速流入", "资金回流"])]
    if not inflow.empty:
        top_in = inflow.nlargest(5, "trend_score")
        names = "、".join(top_in["board_name"].tolist())
        lines.append(f"- 近5日资金流入增强方向：{names}")

    high_trend = df[df["trend_score"] >= 60]
    if not high_trend.empty:
        names_10 = "、".join(high_trend.nlargest(5, "trend_score")["board_name"].tolist())
        lines.append(f"- 近10日主线延续方向：{names_10}")

    divergence = df[df["flow_status"] == "高位分歧"]
    if not divergence.empty:
        names_d = "、".join(divergence["board_name"].tolist()[:3])
        lines.append(f"- 当前风险：{names_d} 高位分歧")
    lines.append("")

    # 二、板块生命周期概览
    lines.append("## 二、板块生命周期概览")
    lines.append("")
    lines.append("| 阶段 | 板块数量 | 代表板块 |")
    lines.append("|---|---:|---|")
    for stage in ["主升期", "加速期", "启动期", "分歧期", "修复期", "退潮期", "沉寂期"]:
        sub = df[df["life_cycle"] == stage]
        cnt = len(sub)
        examples = "、".join(sub.nlargest(3, "trend_score")["board_name"].tolist()) if cnt > 0 else "-"
        lines.append(f"| {stage} | {cnt} | {examples} |")
    lines.append("")

    # 三、状态迁移提醒
    lines.append("## 三、状态迁移提醒")
    lines.append("")
    strengthen = df[df["life_cycle_signal"].isin([
        "新启动", "主线加强", "主线确认", "修复转强", "资金回流加强", "主线延续", "状态转强"
    ])]
    weaken = df[df["life_cycle_signal"].isin([
        "主线分歧", "加速后分歧", "分歧转退潮", "主线退潮", "启动失败", "修复失败", "状态转弱"
    ])]

    if not strengthen.empty:
        lines.append("### 主线加强")
        lines.append("")
        lines.append("| 板块 | 类型 | 昨日 | 今日 | 信号 | 趋势评分 |")
        lines.append("|---|---|---|---|---:|---:|")
        for _, r in strengthen.head(10).iterrows():
            lines.append(f"| {r['board_name']} | {r['board_type']} | {r['prev_life_cycle']} | {r['life_cycle']} | {r['life_cycle_signal']} | {int(r['trend_score'])} |")
        lines.append("")

    if not weaken.empty:
        lines.append("### 风险转弱")
        lines.append("")
        lines.append("| 板块 | 类型 | 昨日 | 今日 | 信号 | 趋势评分 |")
        lines.append("|---|---|---|---|---:|---:|")
        for _, r in weaken.head(10).iterrows():
            lines.append(f"| {r['board_name']} | {r['board_type']} | {r['prev_life_cycle']} | {r['life_cycle']} | {r['life_cycle_signal']} | {int(r['trend_score'])} |")
        lines.append("")

    # 四、近5日资金流向
    lines.append("## 四、近5日资金流向")
    lines.append("")
    lines.append("### 资金流入增强 TOP10")
    lines.append("")
    lines.append("| 板块 | 类型 | 资金状态 | 趋势评分 | 涨幅 | 成交占比 | 5日变化 | 连升 | 领涨股 |")
    lines.append("|---|---|---|---:|---:|---:|---:|---:|")
    inflow5 = df[df["flow_status"].isin(["持续流入", "加速流入", "资金回流"])].nlargest(10, "trend_score")
    for _, r in inflow5.iterrows():
        c5 = _fmt_change_pct(r.get("amount_ratio_change_5d"))
        ar = _fmt_ratio_pct(r.get("latest_amount_ratio"))
        pct = _fmt_num(r.get("latest_pct_chg"), 1)
        streak = int(r.get("amount_ratio_up_streak", 0))
        leader = r.get("leader_name", "-")
        lines.append(
            f"| {r['board_name']} | {r['board_type']} | {r['flow_status']} | "
            f"{int(r['trend_score'])} | {pct}% | {ar} | {c5} | {streak} | {leader} |"
        )
    lines.append("")

    lines.append("### 资金退潮 TOP10")
    lines.append("")
    lines.append("| 板块 | 类型 | 趋势评分 | 涨幅 | 5日变化 |")
    lines.append("|---|---|---:|---:|---:|")
    outflow = df[df["flow_status"] == "资金退潮"].nlargest(10, "trend_score")
    for _, r in outflow.iterrows():
        c5 = _fmt_change_pct(r.get("amount_ratio_change_5d"))
        pct = _fmt_num(r.get("latest_pct_chg"), 1)
        lines.append(f"| {r['board_name']} | {r['board_type']} | {int(r['trend_score'])} | {pct}% | {c5} |")
    lines.append("")

    # 五、近10日主线
    lines.append("## 五、近10日主线延续")
    lines.append("")
    if windows_available < 10:
        lines.append(f"> 历史数据不足 10 个交易日（当前 {windows_available}），暂不生成近10日主线延续。")
        lines.append("")
    lines.append("| 板块 | 类型 | 趋势评分 | 10日变化 | 10日上涨天 | 说明 |")
    lines.append("|---|---|---:|---:|---:|---|")
    top10 = df.nlargest(10, "trend_score")
    for _, r in top10.iterrows():
        c10 = _fmt_change_pct(r.get("amount_ratio_change_10d"))
        p10 = r.get("pct_positive_days_10d") or 0
        lines.append(
            f"| {r['board_name']} | {r['board_type']} | {int(r['trend_score'])} | "
            f"{c10} | {p10} | {_score_to_label(r['trend_score'])} |"
        )
    lines.append("")

    # 六、市场风格
    lines.append("## 六、近20日市场风格")
    lines.append("")
    if windows_available < 20:
        lines.append(f"> 历史数据不足 20 个交易日（当前 {windows_available}），暂不生成近20日市场风格。")
        lines.append("")
    lines.append("| 板块 | 类型 | 趋势评分 | 20日变化 | 状态 |")
    lines.append("|---|---|---:|---:|---|")
    top20 = df.nlargest(10, "trend_score")
    for _, r in top20.iterrows():
        c20 = _fmt_change_pct(r.get("amount_ratio_change_20d"))
        lines.append(
            f"| {r['board_name']} | {r['board_type']} | {int(r['trend_score'])} | "
            f"{c20} | {_score_to_label(r['trend_score'])} |"
        )
    lines.append("")

    # 七、高位分歧
    lines.append("## 七、高位分歧方向")
    lines.append("")
    if not divergence.empty:
        lines.append("| 板块 | 类型 | 成交占比 | 放量 | 涨幅 | 上涨比 | 风险 |")
        lines.append("|---|---|---:|---:|---:|---:|---|")
        for _, r in divergence.iterrows():
            ar = _fmt_ratio_pct(r.get("latest_amount_ratio"))
            vs5 = _fmt_num(r.get("amount_vs_5d_mean") or 1, 1)
            pct = _fmt_num(r.get("latest_pct_chg"), 1)
            ur = f"{r['up_ratio']*100:.0f}%" if r.get("up_ratio") else "-"
            lines.append(
                f"| {r['board_name']} | {r['board_type']} | "
                f"{ar} | {vs5}x | {pct}% | {ur} | 放量不涨，警惕分歧 |"
            )
    else:
        lines.append("暂无高位分歧板块")
    lines.append("")

    # 八、资金回流
    lines.append("## 八、资金回流方向")
    lines.append("")
    backflow = df[df["flow_status"] == "资金回流"]
    if not backflow.empty:
        lines.append("| 板块 | 类型 | 趋势评分 | 5日变化 | 10日变化 | 涨幅 |")
        lines.append("|---|---|---:|---:|---:|---:|")
        for _, r in backflow.iterrows():
            c5 = _fmt_change_pct(r.get("amount_ratio_change_5d"))
            c10 = _fmt_change_pct(r.get("amount_ratio_change_10d"))
            pct = _fmt_num(r.get("latest_pct_chg"), 1)
            lines.append(
                f"| {r['board_name']} | {r['board_type']} | {int(r['trend_score'])} | "
                f"{c5} | {c10} | {pct}% |"
            )
    else:
        lines.append("暂无资金回流板块")
    lines.append("")

    # 九、观察重点
    lines.append("## 九、明日观察重点")
    lines.append("")
    if not inflow.empty:
        top2 = inflow.nlargest(2, "trend_score")
        for _, r in top2.iterrows():
            lines.append(f"- 如果 **{r['board_name']}** 继续成交占比提升，主线确认度提高；")
    if not divergence.empty:
        d = divergence.iloc[0]
        lines.append(f"- 如果 **{d['board_name']}** 放量但上涨比例下降，警惕分歧；")
    if not outflow.empty:
        o = outflow.iloc[0]
        lines.append(f"- 如果 **{o['board_name']}** 成交占比继续下降，降低关注。")
    lines.append("")
    lines.append(f"> 数据覆盖 {len(df)} 个板块，基于 {windows_available} 个交易日数据。")

    return "\n".join(lines)


def _generate_excel(df, trade_date):
    """生成 Excel 追踪表"""
    wb = Workbook()

    for board_type, sheet_name in [("行业", "行业趋势追踪"), ("概念", "概念趋势追踪")]:
        ws = wb.create_sheet(sheet_name) if sheet_name != "行业趋势追踪" else wb.active
        ws.title = sheet_name
        bdf = df[df["board_type"] == board_type].sort_values("trend_score", ascending=False)

        headers = ["板块", "生命周期", "昨日阶段", "阶段变化", "资金状态", "趋势评分",
                   "成交占比", "5日变化", "10日变化", "20日变化", "连续上升", "主线连续天数",
                   "最新涨幅", "领涨股", "领涨涨幅"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for row_idx, (_, r) in enumerate(bdf.iterrows(), 2):
            vals = [
                r["board_name"],
                r.get("life_cycle", ""),
                r.get("prev_life_cycle", ""),
                r.get("life_cycle_signal", ""),
                r.get("flow_status", ""),
                int(r["trend_score"]),
                _excel_num(r.get("latest_amount_ratio"), 100),
                _excel_num(r.get("amount_ratio_change_5d"), 100),
                _excel_num(r.get("amount_ratio_change_10d"), 100),
                _excel_num(r.get("amount_ratio_change_20d"), 100),
                int(r.get("amount_ratio_up_streak", 0)),
                int(r.get("mainline_streak", 0)),
                _excel_num(r.get("latest_pct_chg")),
                r.get("leader_name", ""),
                _excel_num(r.get("leader_pct_chg")),
            ]
            for col, v in enumerate(vals, 1):
                ws.cell(row=row_idx, column=col, value=v)

        ws.freeze_panes = "A2"

    # 主线迁移看板
    ws3 = wb.create_sheet("主线迁移看板")
    categories = [
        ("资金流入增强", ["持续流入", "加速流入", "资金回流"]),
        ("主线延续", []),
        ("高位分歧", ["高位分歧"]),
        ("资金退潮", ["资金退潮"]),
    ]
    row = 1
    for title, statuses in categories:
        ws3.cell(row=row, column=1, value=title)
        row += 1
        ws3.cell(row=row, column=1, value="板块")
        ws3.cell(row=row, column=2, value="类型")
        ws3.cell(row=row, column=3, value="趋势评分")
        row += 1
        if statuses:
            sub = df[df["flow_status"].isin(statuses)]
        elif title == "主线延续":
            sub = df[df["trend_score"] >= 60]
        else:
            sub = pd.DataFrame()
        for _, r in sub.head(10).iterrows():
            ws3.cell(row=row, column=1, value=r["board_name"])
            ws3.cell(row=row, column=2, value=r["board_type"])
            ws3.cell(row=row, column=3, value=int(r["trend_score"]))
            row += 1
        row += 1

    # 指标说明
    ws4 = wb.create_sheet("指标说明")
    notes = [
        ("成交占比", "板块成交额 / 全市场成交额"),
        ("成交占比变化", "最新成交占比 - N日前成交占比"),
        ("持续流入", "成交占比连续提升 + 板块上涨 + 上涨比例≥55%"),
        ("加速流入", "成交占比快速提升 + 成交额明显放大 + 涨幅>1%"),
        ("资金回流", "近5日转正但近10日仍为负 + 上涨"),
        ("高位分歧", "成交活跃但涨幅或上涨比例不强，资金分歧加大"),
        ("资金退潮", "成交占比下降 + 价格转弱"),
        ("趋势评分", "资金(40)+价格(20)+扩散(20)+活跃(20)，满分100"),
    ]
    for i, (k, v) in enumerate(notes, 1):
        ws4.cell(row=i, column=1, value=k)
        ws4.cell(row=i, column=2, value=v)

    path = REPORTS_DIR / f"board_trend_tracker_{trade_date}.xlsx"
    wb.save(path)
    print(f"Excel 已保存：{path}")


def _generate_summary_json(df, trade_date):
    """生成摘要 JSON"""
    inflow = df[df["flow_status"].isin(["持续流入", "加速流入", "资金回流"])].nlargest(10, "trend_score")
    # 生命周期概览
    life_cycle_overview = {}
    for stage in ["主升期", "加速期", "启动期", "分歧期", "修复期", "退潮期", "沉寂期"]:
        sub = df[df["life_cycle"] == stage]
        life_cycle_overview[stage] = sub.nlargest(5, "trend_score")["board_name"].tolist() if not sub.empty else []

    st_signals = ["新启动", "主线加强", "主线确认", "修复转强", "资金回流加强", "主线延续", "状态转强"]
    wk_signals = ["主线分歧", "加速后分歧", "分歧转退潮", "主线退潮", "启动失败", "修复失败", "状态转弱"]
    sdf = df[df["life_cycle_signal"].isin(st_signals)]
    wdf = df[df["life_cycle_signal"].isin(wk_signals)]

    summary = {
        "trade_date": trade_date,
        "life_cycle_overview": life_cycle_overview,
        "strengthening_boards": [
            {"board_name": r["board_name"], "board_type": r["board_type"],
             "prev_life_cycle": r["prev_life_cycle"], "life_cycle": r["life_cycle"],
             "life_cycle_signal": r["life_cycle_signal"], "trend_score": int(r["trend_score"])}
            for _, r in sdf.iterrows()
        ],
        "weakening_boards": [
            {"board_name": r["board_name"], "board_type": r["board_type"],
             "prev_life_cycle": r["prev_life_cycle"], "life_cycle": r["life_cycle"],
             "life_cycle_signal": r["life_cycle_signal"], "trend_score": int(r["trend_score"])}
            for _, r in wdf.iterrows()
        ],
        "top_inflow_5d": [
            {
                "board_name": r["board_name"],
                "board_type": r["board_type"],
                "flow_status": r["flow_status"],
                "trend_score": int(r["trend_score"]),
                "amount_ratio_change_5d": r.get("amount_ratio_change_5d"),
            }
            for _, r in inflow.iterrows()
        ],
        "top_trend_10d": [
            {
                "board_name": r["board_name"],
                "board_type": r["board_type"],
                "trend_score": int(r["trend_score"]),
            }
            for _, r in df.nlargest(10, "trend_score").iterrows()
        ],
        "divergence_boards": [
            {"board_name": r["board_name"], "board_type": r["board_type"]}
            for _, r in df[df["flow_status"] == "高位分歧"].head(5).iterrows()
        ],
        "outflow_boards": [
            {"board_name": r["board_name"], "board_type": r["board_type"]}
            for _, r in df[df["flow_status"] == "资金退潮"].head(5).iterrows()
        ],
        "watch_points": _gen_watch_points(df),
    }
    path = REPORTS_DIR / f"board_trend_summary_{trade_date}.json"
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"JSON 摘要已保存：{path}")
    return summary


def _gen_watch_points(df):
    points = []
    inflow = df[df["flow_status"].isin(["持续流入", "加速流入"])]
    if not inflow.empty:
        top = inflow.nlargest(2, "trend_score")
        for _, r in top.iterrows():
            points.append(f"如果 {r['board_name']} 继续成交占比提升，主线确认度提高")
    divergence = df[df["flow_status"] == "高位分歧"]
    if not divergence.empty:
        points.append(f"如果 {divergence.iloc[0]['board_name']} 放量但上涨比例下降，警惕分歧")
    outflow = df[df["flow_status"] == "资金退潮"]
    if not outflow.empty:
        points.append(f"如果 {outflow.iloc[0]['board_name']} 成交占比继续下降，降低关注")
    return points


def run(trade_date=None, windows=None):
    if trade_date is None:
        trade_date = datetime.now().strftime("%Y%m%d")

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # 读取数据
    windows = windows or [5, 10, 20]
    df = _read_board_data(trade_date)
    if df is None:
        print(f"[警告] board_amount_ratio 无数据或数据库不可用，跳过趋势追踪")
        return

    # 板块名称标准化（Ⅱ/Ⅲ去重 → display_name）
    df = aggregate_by_display_name(df)

    trade_days = sorted(df["trade_date"].dt.strftime("%Y%m%d").unique())
    windows_available = len(trade_days)

    if windows_available < 5:
        print(f"[警告] 历史数据不足 5 个交易日（当前 {windows_available}），暂不生成趋势追踪")
        return

    # 计算指标
    metrics_df = _calc_metrics(df, trade_date, windows)
    if metrics_df.empty:
        print("[警告] 指标计算结果为空")
        return

    # 分类 + 评分 + 生命周期
    metrics_df["flow_status"] = metrics_df.apply(_classify_flow, axis=1)
    metrics_df["trend_score"] = metrics_df.apply(_calc_trend_score, axis=1)
    metrics_df["trend_label"] = metrics_df["trend_score"].apply(_score_to_label)
    metrics_df["life_cycle"] = metrics_df.apply(_classify_lifecycle, axis=1)
    metrics_df["life_cycle_desc"] = metrics_df["life_cycle"].apply(_lifecycle_desc)

    # 状态迁移：计算上一交易日
    prev_trade_dates = sorted([d for d in trade_days if d < trade_date])
    prev_map = {}
    if prev_trade_dates:
        prev_date = prev_trade_dates[-1]
        prev_df = _calc_metrics(df[df["trade_date"] <= datetime.strptime(prev_date, "%Y%m%d")], prev_date, windows)
        if not prev_df.empty:
            prev_df["flow_status"] = prev_df.apply(_classify_flow, axis=1)
            prev_df["trend_score"] = prev_df.apply(_calc_trend_score, axis=1)
            prev_df["life_cycle"] = prev_df.apply(_classify_lifecycle, axis=1)
            prev_map = {(r["board_type"], r["board_name"]): r["life_cycle"] for _, r in prev_df.iterrows()}

    metrics_df["prev_life_cycle"] = metrics_df.apply(
        lambda r: prev_map.get((r["board_type"], r["board_name"]), ""), axis=1
    )
    metrics_df["life_cycle_change"] = metrics_df.apply(
        lambda r: f"{r['prev_life_cycle']} → {r['life_cycle']}" if r["prev_life_cycle"] else r["life_cycle"], axis=1
    )
    metrics_df["life_cycle_signal"] = metrics_df.apply(
        lambda r: _classify_lifecycle_signal(r["prev_life_cycle"], r["life_cycle"]), axis=1
    )
    metrics_df["mainline_streak"] = metrics_df.apply(_calc_mainline_streak, axis=1)

    # 输出
    md = _generate_markdown(metrics_df, trade_date, windows_available)
    md_path = REPORTS_DIR / f"board_trend_report_{trade_date}.md"
    md_path.write_text(md, encoding="utf-8")
    print(f"Markdown 报告已保存：{md_path}")

    _generate_excel(metrics_df, trade_date)
    _generate_summary_json(metrics_df, trade_date)

    # 统计
    inflow_count = len(metrics_df[metrics_df["flow_status"].isin(["持续流入", "加速流入"])])
    high_trend = (metrics_df["trend_score"] >= 60).sum()
    print(f"趋势追踪完成：{len(metrics_df)} 个板块，流入 {inflow_count}，主线 {high_trend}")


def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
    parser = argparse.ArgumentParser(description="板块资金趋势追踪")
    parser.add_argument("--date", type=str, default=None, help="日期 YYYYMMDD")
    parser.add_argument("--windows", type=int, nargs="*", default=[5, 10, 20], help="窗口天数")
    args = parser.parse_args()
    run(trade_date=args.date, windows=args.windows)


if __name__ == "__main__":
    main()
